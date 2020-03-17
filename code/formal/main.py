import urllib.request
import json
from openpyxl import Workbook,load_workbook


# 获取数据
def get_json(url,file_json):

    f = open(file_json,'w',encoding = 'utf-8')
    r = urllib.request.urlopen(url)
    j = r.read().decode('utf-8')
    print(type(j))
    f.write(j)
    f.close()
    print("**********数据获取成功**********")
    return file_json

#处理数据，获取到的是json源文件，需要解析出其中我们需要的数据并存入txt文件
def deal_json(txt_file,json_file):
    namelist = []
    numlist = []
    with open(json_file,'r',encoding='utf-8') as f:
        jsons = json.loads(f.read())
        List = jsons["result"]['list']
        #print(type(List))
        #print(List)
        fs = open(txt_file,'w',encoding = 'utf-8')
        for i in List:
            namelist.append(i['seriesname'])
            numlist.append(i['righttitle'])

        for r,s in zip(namelist,numlist):
            pat = r.replace(' ', '') + ' ' + s + '\n'
            fs.write(pat)
        fs.close()
    print('**********数据转换成功**********')
    return txt_file

# 存储数据
# 文本读取函数，从解析的得到txt文件中读取数据，并存储为元组
def readTxt(file):
    ls = list()
    with open(file, 'r', encoding='utf-8-sig') as f:
        for line in f.readlines():
            # 异常处理机制
            try:
                index1 = line.index(' ')
                index2 = line.index('\n')
                car_name = line[:index1]
                sale = line[index1 + 1:index2]

                ls.append((car_name, sale))  # 以元组的形式追加进空列表
            except:
                print('wrong format!')
    # 返回一个列表
    return ls

# 数据写入函数（写入excle表格)
def write_excel_xlsx1(path, sheet_name, value):
    index = len(value)  # 列表中所含元组的个数，从而确定写入Excel的行数
    # 打开Excel
    wb = load_workbook(path)
    # 创建工作簿
    sheet = wb.create_sheet(sheet_name)
    # 设置格式(列宽)
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 15
    # 按行加入
    for i in range(index):
        sheet.append(value[i])
    # 保存文件
    wb.save(path)
    print("**********数据写入成功**********")

#创建excle表格，并写入数据
def create_xls(xsl_file,txt_file):
    data_xlsx = xsl_file  # 文件路径，根据自己的需要自行修改
    wb = Workbook()
    wb.save(data_xlsx)
    # 工作簿名称
    sheet_name_xlsx = 'Sales'
    art = readTxt(txt_file)  # 调用读取函数
    # 插入表头
    art.insert(0, ('car_name','Sales_volume'))
    write_excel_xlsx1(data_xlsx, sheet_name_xlsx, art)  # 调用写入函数

if __name__ == "__main__":
    data = r'data.txt'
    excle = 'data.xlsx'
    url = 'https://cars.app.autohome.com.cn/cars_v9.1.0/cars/getseriesranklist.ashx?pageindex=1&pm=1&pluginversion=10.5.5&typeid=1&data=2020-01&cityid=341000&provinceid=340000'
    res = deal_json(data,get_json(url,'text.json'))
    create_xls(excle,data)
